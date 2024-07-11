"""
# Example usage with two variables
variable_value_1 = 42
variable_value_2 = "Hello"
debug_print("Debug message with two variables: {} and {}", variable_value_1, variable_value_2)
"""
####################   snippet for create switch button ########################
#  Tempalate for usage 
"""
#def open_sub_window():
#    sub_window = tk.Toplevel(root)
#    app = SwitchApp(sub_window)

    # Add buttons dynamically to the subwindow
    app.add_button("EngMode", app.img_eng_mode_on, app.img_eng_mode_off)
    app.add_button("VppMode", app.img_vpp_on, app.img_vpp_off)
"""
########################################################################################
#  color : magenta,  cyan,   
NAVY_BLUE = '#868d91'
CADET_BLUE = '#5f9ea0'
SAXE_BLUE ='#4798b3'
BROWN = '#d08f34'
MALACHITE_GREEN = '#00a15c'
ORANGE = '#ea7343'
BLUE_1 = "#{:02X}{:02X}{:02X}".format(112, 114, 209)
ORANGE_1 = "#{:02X}{:02X}{:02X}".format(237, 149, 84)
########################################################################################

from tkinter import Frame, StringVar, Button 
from tkinter import Label, Text, END, W, E, RIGHT, LEFT, ttk, Tk, TOP, BOTTOM
from tkinter import Scrollbar, WORD , Y, BOTH, Entry
from tkinter import colorchooser, IntVar, NORMAL, GROOVE, DISABLED
from .lib_1 import show_warning_message, crc8
from .layout import update_dc_write_color
from .lib_1 import debug_print, is_valid_hex
import threading
import os
from tkinter import PhotoImage
import time
import matplotlib.pyplot as plt
import openpyxl
from openpyxl import Workbook, load_workbook

#from .effect_button import SwitchApp
import serial  # Make sure to import the 'serial' module if it's not already imported
from tkinter import Radiobutton, Checkbutton
import json
from tkinter import filedialog
file_content = ""
from .common import DIM_ID, READ_NUM , START_REG_ADDR , READ_ROW_NUM, READ_NUM_INT
from .items import ITEMS_INA228,  ITEMS_OTP, ITEMS_DB, ITEMS_DC, ITEMS_SETTINGS, ITEMS_VOL_CUR
from .items import get_row_number
ORANGE = '#ea7343'
new_line = False
HT2_NUM = 1

def cmd_to_ser(new_line,main_gui_instance, new_text):
    if (new_line):
        main_gui_instance.set_input_text_continuous(new_text)
    else: 
        main_gui_instance.set_input_text(new_text)

def cmd_to_ser_continous(main_gui_instance, new_text):
    main_gui_instance.set_input_text_continuous(new_text)

############################################
# Get the filename from the user
def write_content_to_file():
    file_content = get_file_content()
    file_path = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=[("Text files", "*.txt")])
    try:
        with open(file_path, 'w') as file:
            file.write(file_content)
        debug_print(f"Content {file_content} successfully written to {file_path}")
    except Exception as e:
        debug_print(f"Error writing content to {file_path}: {e}")

def load_file(file_path):
    try:
        with open(file_path, 'r') as file:
            content = file.read()
        debug_print(f"Loaded file content: {content}")
        process_loaded_content(content)
    except Exception as e:
        debug_print(f"Error loading file: {e}")

def load_command():
    file_path = filedialog.askopenfilename(defaultextension=".txt", filetypes=[("Text files", "*.txt")])
    if file_path:
        load_file(file_path)    

def get_label_2nd_var_values(window_title):
    label_2nd_var_values = []  # List to store label_2nd_var values

    # Get the total number of rows for the specified window_title
    total_rows = len(label_widgets_dict.get(window_title, {}))

    # Iterate through label_widgets_dict
    for row_index, row_data in label_widgets_dict.get(window_title, {}).items():
        # Check if label_2nd_var exists in row_data
        if row_index < 3 or row_index == total_rows: #ignore row_index 1 and 2
            continue

        if 'label_2nd_var' in row_data:
            label_2nd_var_values.append(row_data['label_2nd_var'].get())

    # Return the concatenated string with spaces
    return ' '.join(label_2nd_var_values)

# Initialize an empty string
def get_file_content():
    global file_content
    return file_content
    # sub.py

def process_loaded_content(content):
    global file_content
    file_content = content
    print(f"File content in sub.py updated: {file_content}")

def extract_color_from_file_content(win_name):
    # Assuming each line in file_content is in the format "item_text value"
    lines = file_content.split('\n')
    for line in lines:
        if line.startswith(win_name):
            _, color = line.split(' ', 1)
            return color
    return None  # Return None if item_text is not found in file_content

def extract_value_from_file_content(item_text, file_content):
    # Assuming each line in file_content is in the format "item_text value"
    lines = file_content.split('\n')
    for line in lines:
        if line.startswith(item_text[0:4]):
            _, value = line.split(' ', 1)
            return value
    return None  # Return None if item_text is not found in file_content

label_widgets_dict = {}
########################################################################################
# Function to change the background color of label widgets
# switch_app_instance.add_switch(row_index, column_index, "SwitchButtonName", class_instance, on_image, off_image, "#ff0000")
def update_label_widgets_bg_color(win_name, color,switch_dict):
    # This will update the background color of all switch buttons associated with "MyWindow" in switch_dict.
    def update_switch_buttons_color(win_name, color, switch_dict):
        window_title = win_name.title()
        if window_title in switch_dict:
            for switch_row_index, switch_data in switch_dict[window_title].items():
                debug_print("change switch color ")
                app_switch_instance = switch_data.get("app_switch_instance")
                if app_switch_instance:
                    for button_name, button_data in app_switch_instance.buttons.items():
                        switch_button = button_data.get("button")
                        if switch_button:
                            switch_button.configure(bg=color)
        else:
            debug_print(f"Window '{window_title}' not found in switch_dict.")
    # End of change switch background color thru switch_dict

    for window_title, rows in label_widgets_dict.items():
        if window_title == win_name.title():
            update_switch_buttons_color(win_name, color, switch_dict) #update switch backgroud color
            for row_index, widgets in rows.items():
                try: 
                # Check if the key exists before configuring
                    if "label_2nd_widget" in widgets and isinstance(widgets["label_2nd_widget"], Label):
                        #debug_print("label_2nd_widget change color")
                        widgets["label_2nd_widget"].configure(bg=color)

                    if "check_button" in widgets and isinstance(widgets["check_button"], Checkbutton):
                        widgets["check_button"].configure(bg=color)

                    if "label_trim_value" in widgets and widgets["label_trim_value"] is not None:
                        label_trim_value = widgets["label_trim_value"]
                        label_trim_value.config(bg=color)

                    if "label_iteration_num" in widgets and widgets["label_iteration_num"] is not None:
                        label_iteration_num = widgets["label_iteration_num"]
                        label_iteration_num.config(bg=color)

                    if "radio_button_frame" in widgets and widgets["radio_button_frame"] is not None:
                        #debug_print(" ")
                        radio_button_frame = widgets["radio_button_frame"]                        
                        radio_button_frame.configure(bg=color)                       
                    if "radio_button" in widgets and widgets["radio_button"] is not None:
                    # Iterate over the list of Radiobuttons and configure each one
                        #debug_print(" ")
                        for radio_button in widgets["radio_button"]:
                            debug_print("Checking Radiobutton {}",radio_button)
                            if isinstance(radio_button, Radiobutton):
                                debug_print("Configuring Radiobutton {}",radio_button)
                                radio_button.configure(bg=color)
                            else:                                
                                debug_print("Radiobutton {} is not an instance of tk.Radiobutton",radio_button)
                except Exception as e:
                    debug_print("configuring widgets: {}",e)
                    # Check if the switch_dict for the current window exists

######################################################################################################
def change_bg_color(win_name, switch_dict):
    global file_content
    color = colorchooser.askcolor()[1]  # Get the selected color
    if color:
        # Change the window's background color
        win_name.configure(bg=color)
        file_content += f"{win_name.title()} {color}\n"

        # Change the background color of label widgets
        update_label_widgets_bg_color(win_name, color, switch_dict)

def update_label_2nd_var(window_title, row_index, item_text, new_text, text_color=None, width=None):
    global label_widgets_dict
    global file_content
    if window_title in label_widgets_dict and row_index in label_widgets_dict[window_title]:
        label_2nd_var = label_widgets_dict[window_title][row_index]["label_2nd_var"]
        label_2nd_var.set(new_text)
        file_content += f"{item_text[0:4]} {new_text}\n"
        label_2nd_widget = label_widgets_dict[window_title][row_index]["label_2nd_widget"]
        label_2nd_widget.config(text=new_text)

        if text_color:
            label_2nd_widget.config(fg=text_color)

        if width:
            label_2nd_widget.config(width=width)
    else:
        pass
        #print(f"Row {row_index} not found for window {window_title}")

def update_label_trim_var(window_title, row_index, new_trim_value, new_iteration_num, text_color=None):
    # Accessing label_trim_value and label_iteration_num in some other part of your code
    # print(f"Debug: label_widgets_dict is {label_widgets_dict}")
    if window_title in label_widgets_dict and row_index in label_widgets_dict[window_title]:
        debug_print(f"Debug: label_trim_value type is:\n")
        #debug_print(f"Debug: label_trim_value type is {type(label_trim_value)}\n")
        label_trim_value = label_widgets_dict[window_title][row_index]["label_trim_value"]
        label_iteration_num = label_widgets_dict[window_title][row_index]["label_iteration_num"]

        if label_trim_value:
            debug_print(f"Debug: label_trim_value type is {type(label_trim_value)}")
        # Now you can use label_trim_value and label_iteration_num as needed
            label_trim_value.config(text=new_trim_value)

        if label_iteration_num:            
            # print(f"Debug: label_iteration_num type is {type(label_iteration_num)}")
            label_iteration_num.config(text=new_iteration_num)

        if text_color and label_trim_value and label_iteration_num:
            # print(f"Debug: 46 lib.py text_color is {text_color}, iteration_num is {new_iteration_num}")
            label_trim_value.config(fg=text_color)
            label_iteration_num.config(fg=text_color)
    else:
        debug_print("Error: row {} not found for window {}",row_index,window_title)


class mainGUIFrame(Frame):
    def __init__(self, master=None, ina228_vshunt_value=None, ina228_current_value=None, **kwargs):
        super().__init__(master, **kwargs)
        self.ina228_vshunt_value = ina228_vshunt_value
        self.ina228_current_value = ina228_current_value
        self.uartState = False # is uart open or not

        self.parts_array = []
        self.excel_file = 'output.xlsx'
        self.workbook, self.sheet = self.load_or_create_workbook()
        # a frame contains COM's information, and start/stop button
        self.frame_COMinf = Frame(self)
        self.frame_COMinf.grid(row = 0, column = 1, columnspan =5,sticky="nw")

        labelCOM = Label(self.frame_COMinf,text="COMx: ")
        self.COM = StringVar(value = "COM6")
        comboPort = ttk.Combobox(self.frame_COMinf, width = 10, textvariable=self.COM)
        comboPort["values"] = ("COM3", "COM4", "COM5", "COM6")
        comboPort["state"] = "readonly"
        labelCOM.grid(row = 0, column = 1, padx = 5, pady = 3)
        comboPort.grid(row = 0, column = 2, padx = 5, pady = 3)

        labelBaudrate = Label(self.frame_COMinf,text="Baudrate: ")
        self.Baudrate = StringVar(value = 115200)
        comboBaudrate = ttk.Combobox(self.frame_COMinf, width = 10, textvariable=self.Baudrate)
        comboBaudrate["values"] = ("2400", "4800", "9600", "19200", "115200", "23040" , "1000000")
        comboBaudrate["state"] = "readonly"
        labelBaudrate.grid(row = 0, column = 3, padx = 5, pady = 3)
        comboBaudrate.grid(row = 0, column = 4, padx = 5, pady = 3)

        labelParity = Label(self.frame_COMinf,text="Parity: ")
        self.Parity = StringVar(value ="NONE")
        comboParity = ttk.Combobox(self.frame_COMinf, width = 10, textvariable=self.Parity)
        comboParity["values"] = ("NONE","ODD","EVEN","MARK","SPACE")
        comboParity["state"] = "readonly"
        labelParity.grid(row = 1, column = 1, padx = 5, pady = 3, sticky="nw")
        comboParity.grid(row = 1, column = 2, padx = 5, pady = 3)

        labelStopbits = Label(self.frame_COMinf,text="Stopbits: ")
        self.Stopbits = StringVar(value ="1")
        comboStopbits = ttk.Combobox(self.frame_COMinf, width = 10, textvariable=self.Stopbits)
        comboStopbits["values"] = ("1","1.5","2")
        comboStopbits["state"] = "readonly"
        labelStopbits.grid(row = 1, column = 3, padx = 5, pady = 3)
        comboStopbits.grid(row = 1, column = 4, padx = 5, pady = 3)
        
        self.buttonSS = Button(self.frame_COMinf, text = "Start", command = self.processButtonSS)
        self.buttonSS.grid(row = 2, column = 4, padx = 5, pady = 3, sticky = E)

        # serial object
        self.ser = serial.Serial()

        # serial read threading
        self.ReadUARTThread = threading.Thread(target=self.ReadUART, daemon=True)
        # or self.ReadUARTTHread.daemon = True
        # self.ReadUARTThread.start()

    ################## CREATE Receive area ######################################

        frameRecv = Frame(self)
        frameRecv.grid(row = 1, column = 1, columnspan=5, sticky = 'nsew')
        labelOutText = Label(frameRecv,text="Received Data:")
        labelOutText.grid(row = 1, column = 1, padx = 3, pady = 2, sticky = W)

        frameRecvSon = Frame(frameRecv)
        frameRecvSon.grid(row = 2, column =1, columnspan=5, sticky='nsew' )

        self.buttonClear = Button(frameRecv, text="Clear", command=self.processButtonClear)
        self.buttonClear.grid( row = 3, column= 1, padx = 5, pady =3, sticky = W)

        scrollbarRecv = Scrollbar(frameRecvSon)
        scrollbarRecv.grid(row=0, column=1, sticky='ns')
        #scrollbarRecv.pack(side = RIGHT, fill = Y)

        self.OutputText = Text(frameRecvSon, wrap = WORD, width = 60, height = 10, yscrollcommand = scrollbarRecv.set)
        self.OutputText.grid(row=0, column=0, sticky='nsew')
        scrollbarRecv.config(command=self.OutputText.yview)

        # Configure grid weights to make the text widget expand
        frameRecv.grid_rowconfigure(2, weight=1)
        frameRecv.grid_columnconfigure(1, weight=1)

        frameRecvSon.grid_rowconfigure(0, weight=1)
        frameRecvSon.grid_columnconfigure(0, weight=1)

    #############################################################################

    ################## CREATE transmit area ######################################

        #frameTrans = Frame(self, height=200)
        frameTrans = Frame(self)
        frameTrans.grid(row = 2, column = 1, columnspan=4, stick= 'w')
        labelInText = Label(frameTrans,text="To Transmit Data:")
        labelInText.grid(row = 1, column = 1, padx = 3, pady = 2, sticky ='w')

        # Save Settings Button
        self.buttonSaveSettings = Button(frameTrans, text="Save Settings", command=write_content_to_file)
        self.buttonSaveSettings.grid(row=1, column=3, padx=5, pady=3, stick='w')

        # Load Settings Button
        self.buttonLoadSettings = Button(frameTrans, text="Load Settings", command=load_command)
        self.buttonLoadSettings.grid(row=1, column=4, padx=5, pady=3, stick='w')

        # Clear Button
        self.buttonClear = Button(frameTrans, text="Clear", command=self.clearInputText)
        self.buttonClear.grid(row=5, column=1, padx=5, pady=3, sticky='w')

        # Save Cmds Button
        self.buttonSaveCmds = Button(frameTrans, text="Save Cmds", command=self.write_cmds_to_file)
        self.buttonSaveCmds.grid(row=5, column=3, padx=5, pady=3, stick='w')

        # Load Cmds Button
        self.buttonLoadCmds = Button(frameTrans, text="Load Cmds", command=self.load_cmds)
        self.buttonLoadCmds.grid(row=5, column=4, padx=5, pady=3)

        frameTransSon = Frame(frameTrans)
        frameTransSon.grid(row = 2, column =1, columnspan = 5, sticky='nsew')

        scrollbarTrans = Scrollbar(frameTransSon)
        scrollbarTrans.grid(row=0, column=1, sticky='ns')
        #scrollbarTrans.pack(side = RIGHT, fill = Y)

        self.InputText = Text(frameTransSon, wrap = WORD, width = 60, height = 20, yscrollcommand = scrollbarTrans.set)
        self.InputText.grid(row=0, column=0, sticky='nsew')

        scrollbarTrans.config(command=self.InputText.yview)

        # Send Button
        self.buttonSend = Button(frameTrans, text = "Send", command = self.processButtonSend)
        self.buttonSend.grid(row = 5, column = 5, padx = 5, pady = 3, sticky = E)

        # Configure grid weights to make the text widget expand
        frameTrans.grid_rowconfigure(2, weight=1)
        frameTrans.grid_columnconfigure(1, weight=1)

        frameTransSon.grid_rowconfigure(0, weight=1)
        frameTransSon.grid_columnconfigure(0, weight=1)

    ################## CREATE transmit area ######################################

         # Configure grid weights for resizing
        self.grid_rowconfigure(0, weight=1)  # Allow row 0 to expand vertically
        self.grid_rowconfigure(1, weight=1)  # Allow row 1 to expand vertically
        self.grid_rowconfigure(2, weight=1)  # Allow row 2 to expand vertically
        self.grid_columnconfigure(0, weight=1)  # Allow column 0 to expand horizontally
        self.grid_columnconfigure(1, weight=1)  # Allow column 1 to expand horizontally
        self.grid_columnconfigure(2, weight=1)  # Allow column 2 to expand horizontally
        self.grid_columnconfigure(3, weight=1)  # Allow column 3 to expand horizontally
        self.grid_columnconfigure(4, weight=1)  # Allow column 4 to expand horizontally

        self.pack(fill=BOTH, expand=True) 

    def load_cmds(self):
        file_path = filedialog.askopenfilename(defaultextension=".txt", filetypes=[("Text files", "*.txt")])
        if file_path:
            try:
                with open(file_path, 'r') as file:
                    content = file.read()
                print(f"Loaded file content: {content}")
                self.InputText.delete("1.0", END)  # Clear existing content in InputText
                self.InputText.insert("1.0", content)  # Insert loaded content into InputText
            except Exception as e:
                print(f"Error loading file: {e}")

    def write_cmds_to_file(self):
        cmds_content = self.InputText.get("1.0", END) 
        file_path = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=[("Text files", "*.txt")])
        try:
            with open(file_path, 'w') as file:
                file.write(cmds_content)
                print(f"Content {cmds_content} successfully written to {file_path}")
        except Exception as e:
            print(f"Error writing content to {file_path}: {e}")

    def processButtonClear(self):
        self.OutputText.delete(1.0, END)

    def clearInputText(self):
        self.InputText.delete(1.0, END)
    def processButtonSS(self):
        print(self.Parity.get())
        if (self.uartState):
            self.ser.close()
            self.buttonSS["text"] = "Start"
            self.uartState = False
            print("here is self.uartState = 1")
        else:
            # restart serial port
            self.ser.port = self.COM.get()
            self.ser.baudrate = self.Baudrate.get()
            print(f"self.ser.baudrate is {self.ser.baudrate}")
            
            strParity = self.Parity.get()
            if (strParity=="NONE"):
                self.ser.parity = serial.PARITY_NONE
            elif(strParity=="ODD"):
                self.ser.parity = serial.PARITY_ODD
            elif(strParity=="EVEN"):
                self.ser.parity = serial.PARITY_EVEN
            elif(strParity=="MARK"):
                self.ser.parity = serial.PARITY_MARK
            elif(strParity=="SPACE"):
                self.ser.parity = serial.PARITY_SPACE
                
            strStopbits = self.Stopbits.get()
            if (strStopbits == "1"):
                self.ser.stopbits = serial.STOPBITS_ONE
            elif (strStopbits == "1.5"):
                self.ser.stopbits = serial.STOPBITS_ONE_POINT_FIVE
            elif (strStopbits == "2"):
                self.ser.stopbits = serial.STOPBITS_TWO
            
            try:
                self.ser.open()
            except:
                infromStr = "Can't open "+self.ser.port
                InformWindow(infromStr)
            
            if (self.ser.isOpen()): # open success
                self.buttonSS["text"] = "Stop"
                debug_print("set serial commuinication")
                self.uartState = True
                self.ReadUARTThread.start()

    def processButtonSend(self):
        try: 
            if (self.uartState):
                strToSend = self.InputText.get(1.0,END)
                #bytesToSend = (strToSend[0:-1] + '\n').encode(encoding='ascii')
                lines = strToSend.splitlines()
                for line in lines:
                    bytesToSend = (line + '\n').encode(encoding='ascii')
                    self.ser.write(bytesToSend)
                    print("Bytes sent:",bytesToSend)
                    time.sleep(1.0)
            else:
                infromStr = "Not In Connect!"
                InformWindow(infromStr)
        except Exception as e:
            print("Error during write operation:", str(e))
            # Handle the error as needed

    def float_to_bit_binary(self, float_value, bit):
    # Convert float to binary string for the integer part
        integer_part = int(float_value)
        int_bit = int(bit)
        fractional_part = float_value - integer_part

        if fractional_part >=0.5:
            integer_part += 1

        integer_binary = bin(integer_part)[2:].zfill(int_bit)  # Convert to binary and pad to three bits
        print(integer_binary)

    # Convert fractional part to binary string for three fractional bits
        fractional_binary = ""
        for _ in range(3):
            fractional_part *= 2
            if fractional_part >= 1:
                fractional_binary += "1"
                fractional_part -= 1
            else:
                fractional_binary += "0"

    # Combine integer and fractional binary strings
        binary_string = integer_binary + fractional_binary
        return integer_binary

    def load_or_create_workbook(self):
        if os.path.exists(self.excel_file):
            workbook = load_workbook(self.excel_file)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
        return workbook, sheet

    def ReadUART(self):
        # RB ROW_NUM XX READ_NUM YY u32 u32 u32
        global START_REG_ADDR
        global READ_ROW_NUM
        global READ_NUM_INT
        global HT2_NUM

        coarse_freq_time = 1
        fine_freq_time = 1
        debug_print(f"Threadng .... \n")
        while self.uartState: #True
            if (self.uartState):
                try:
                    data = self.ser.readline().decode(encoding='ascii').strip()
                    debug_print(f"Received data: '{data}'")
                    #print( data,end='')
                    #data_without_newline = data.replace('\n', '').replace('\r', '')
                    #self.OutputText.insert(END, data.replace('\n', '') + '\n')
                    parts = data.split()
                    if parts[0][:2] == "RB":
                        ROW_NUM = int(parts[2])
                        READ_NUM = int(parts[4]) # Integer division to ensure it is always an integer 
                        result_0 = self.split_string_into_chunks(parts[5], 2)
                        debug_print(f"READ_NUM is: {READ_NUM}\n")
                        for i in range(READ_NUM):
                            if parts[0][2] == "C":
                                update_label_2nd_var("win_dc_wr_one", ROW_NUM + i, "", result_0[i+1], "dark red" )
                            elif parts[0][2] == "B":
                                update_label_2nd_var("win_db_wr", ROW_NUM + i, "", result_0[i+1], "dark red" )
                                # debug_print(f"result_0 is: {result_0[i]}, i : {i}\n")
                            elif parts[0][2] == "T":
                                update_label_2nd_var("win_otp", ROW_NUM + i, "", result_0[i+1], "dark red" )

                    elif parts[0] == "Freq" and len(parts) == 3:
                        debug_print(f"received freq : {parts}\n")
                        string1, string2, string3 = parts
                        if ( not string3 == "0.000000" ) :
                            freq_value = float(string3)
                            freq_text = "{:.3f}".format(freq_value) # always make sure two digits before the decimal point and three digits after.

                            match string2:
                                case 'coarse':
                                    code_float = 3.0 + (10.0 * ((62.9/freq_value) -1.0))
                                    code_binary = self.float_to_bit_binary(code_float, 3) # only 3 bits
                                    #debug_print(f"code for float : {code_float}\n")
                                    row_index = get_row_number(ITEMS_VOL_CUR, "0xc9 COARSE FREQ") + 1 # plus one by OP08
                                    #print(freq_value)
                                    update_label_trim_var("win_trim", row_index, freq_text, coarse_freq_time, 'blue')
                                    update_label_2nd_var("win_trim", row_index, "", code_binary , "dark red", "" )
                                    coarse_freq_time += 1
                                case 'fine':
                                    #code_float = 16.0 - ((freq_value - 62.9)/(freq_value * 0.00478))
                                    code_float = 16.0 - ((freq_value - 62.9)/(freq_value * 0.00478))
                                    code_binary = self.float_to_bit_binary(code_float, 5) # only 5 bits 
                                    row_index = get_row_number(ITEMS_VOL_CUR, "0xcb FINE FREQ") + 1 # plus one by OP08
                                    update_label_trim_var("win_trim", row_index, freq_text, 1, 'blue')
                                    update_label_2nd_var("win_trim", row_index, "", code_binary , "dark red", "" )
                                    fine_freq_time += 1
                        elif ( string3 == "0.000000" ) :
                            self.OutputText.insert(END,f"ERROR : Freq measurement fail: {string3}Mhz\n", 'error_color')
                            self.OutputText.tag_config('error_color', foreground='red')

                    elif parts[0] == "VBUS" and len(parts) == 4:
                        string1, string2, string3, string4 = parts
                        if string4 == "time":
                                row_index = get_row_number(ITEMS_VOL_CUR, "0xcd VOLTAGE") # plus one by OP08
                                #debug_print(f"String 1 : {string1}, String 2: {string2}, String 3: {string3}, String 4: {string4}, row_index : {row_index}\n")
                                update_label_trim_var("win_trim", row_index, string2, string3, 'blue')
                    elif parts[0] == "VAVDDL" and len(parts) == 3:
                        string1, string2, string3 = parts
                        debug_print(f"String 1 : {string1}, String 2: {string2}, String 3: {string3}\n")
                        if string2 == "Code":
                                row_index = get_row_number(ITEMS_VOL_CUR, "0xcd VOLTAGE") # plus one by OP08
                                update_label_2nd_var("win_trim", row_index, "", string3 , "dark red", "" )

                                # to update otp VREF value
                                row_index = get_row_number(ITEMS_OTP, "0x65 VREF_TRIM[3:0]") # plus one by OP08
                                current_trim_code = int(string3, 2)
                                hex_value_current_trim_code = hex(current_trim_code)
                                string_current_code = str(hex_value_current_trim_code)[2:]
                                debug_print(f"String 3 : {string3}, current_trim_code : {current_trim_code}, hex_value: {hex_value_current_trim_code}, string : {string_current_code}\n")
                                update_label_2nd_var("win_otp", row_index, "", string_current_code , "dark red", "" )

                    elif parts[0] == "INA228" and len(parts) == 3:
                        string1, string2, string3 = parts
                        debug_print(f"String 1 : {string1}, String 2: {string2}, String 3: {string3}\n")
                        self.OutputText.insert(END,f"{string1} {string2} {string3}\n")

                        if string1 == "INA228" and string2 == "VSHUNT" :
                            text_to_insert = f"{string3} V"
                            self.ina228_vshunt_value.insert("1.0", text_to_insert, "right") 

                        if string1 == "INA228" and string2 == "CURRENT_in_Amp" :
                            text_to_insert = f"{string3} A"
                            self.ina228_current_value.insert("1.0", text_to_insert, "right") 
                            debug_print(f"String 1 : {string1}, String 2: {string2}, String 3: {string3}\n")
                    elif parts[0] == "PASS":
                        self.OutputText.insert(END,f"{data}\n", 'PASS_color')
                        self.OutputText.tag_config('PASS_color', foreground='blue')
                    elif parts[0] == "FAIL":
                        self.OutputText.insert(END,f"{data}\n", 'FAIL_color')
                        self.OutputText.tag_config('FAIL_color', foreground='red')
                    elif parts[0] == "WARNING":
                        self.OutputText.insert(END,f"{data}\n", 'WARNING_color')
                        self.OutputText.tag_config('WARNING_color', foreground='green')
                    elif parts[0] == "RUNNING":
                        self.OutputText.insert(END,f"{data}\n", 'PASS_color')
                        self.OutputText.tag_config('PASS_color', foreground='blue')
                    elif parts[0] == "LED_CURRENT":
                        self.OutputText.insert(END,f"{data}\n", 'PASS_color')
                        self.OutputText.tag_config('PASS_color', foreground='blue')
                        if parts[1] == "measure_over":
                            file_path = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=[("Text files", "*.txt")])
                            try:
                                with open(file_path, 'w') as file:
                                    # Ensure self.parts_array is a string
                                    content_to_write = str(self.parts_array)
                                    file.write(content_to_write)
                                    print(f"Content successfully written to {file_path}")
                                    self.parts_array = []
                            except Exception as e:
                                print(f"Error writing content to {file_path}: {e}")
                        else: 
                            self.parts_array.append((parts[4], float(parts[5])))
                            debug_print(f"Updated parts_array: {self.parts_array}")

                            row = int(parts[4]) + 1
                            value = float(parts[5])
                            column = int(parts[2])
                            self.save_to_excel(row, value, column)
                            #if int(parts[4]) == HT2_NUM:
                            #    self.plot_parts_array()

                except UnicodeDecodeError as e:
                    self.handle_exception(f"Error decoding received data: {e}")
                except serial.SerialException as e:
                    self.handle_exception(f"Serial communication error: {e}")
                except Exception as e:
                    self.handle_exception(f"Something went wrong in receiving: {e}")

                # Adjust the polling interval as needed                    
                time.sleep(0.10)                    

        # Ensure proper cleanup when the loop exits
        self.ser.close()
        self.buttonSS["text"] = "Start"               

    def plot_parts_array(self):
        # Extract x and y values from parts_array
        x_values = [int(pair[0]) for pair in self.parts_array]
        y_values = [float(pair[1]) for pair in self.parts_array]

        # Plotting the values
        plt.figure()
        plt.plot(x_values, y_values, marker='o')
        plt.xlabel('X Axis (parts[1])')
        plt.ylabel('Y Axis (parts[2])')
        plt.title('LED Current Data')
        plt.grid(True)
        plt.show()

        # Reset the parts_array after plotting
        self.parts_array = []

    def save_to_excel(self, row, value, column):
        col_num = int(column)
        row_num = int(row)
        self.sheet.cell(row=row_num, column=col_num, value=value)
        self.workbook.save(self.excel_file)

    def split_string_into_chunks(self,input_string, chunk_size=2):
        return [input_string[i:i+chunk_size] for i in range(0, len(input_string), chunk_size)]

    def handle_exception(self, error_message):
        inform_str = f"Error: {error_message}"
        InformWindow(inform_str)
        self.ser.close()
        self.buttonSS["text"] = "Start"
        self.uartState = False

    @staticmethod
    def rgb_to_hex(r,g,b):
        return f'#{r:02x}{g:02x}{b:02x}'

    def set_input_text(self, text):
        self.InputText.delete(1.0,END)                    
        words = text.split()
        match words[0]:
            case 'SW':
                color_tag = 'SW_color'
            case 'SET':
                color_tag = 'SET_color'            
            case 'CW':
                #color_tag = 'purple_color'            
                color_tag = 'CW_color'            
            case  'BW':
                color_tag = 'BW_color'            
            case _:
                color_tag = 'black_color'
            
        self.InputText.insert(END, text, color_tag)
        self.InputText.tag_config('SW_color', foreground='green')
        self.InputText.tag_config('black_color', foreground='black')
        self.InputText.tag_config('SET_color', foreground='#8B4513')  # Brown color code
        self.InputText.tag_config('purple_color', foreground='#800080')  # Brown color code
        self.InputText.tag_config('CW_color', foreground='blue')  # Brown color code
        self.InputText.tag_config('BW_color', foreground=self.rgb_to_hex(241,120,80))  # call rgb_to_hex by self

    def set_input_text_continuous(self, text):
        text = '\n' + text
        words = text.split()
        match words[0]:
            case 'SW':
                color_tag = 'SW_color'
            case 'SET':
                color_tag = 'SET_color'            
            case 'CW':
                #color_tag = 'purple_color'            
                color_tag = 'CW_color'            
            case  'BW':
                color_tag = 'BW_color'            
            case _:
                color_tag = 'black_color'
            
        self.InputText.insert(END, text, color_tag)
        self.InputText.tag_config('green_color', foreground='green')
        self.InputText.tag_config('black_color', foreground='black')
        self.InputText.tag_config('brown_color', foreground='#8B4513')  # Brown color code
        self.InputText.tag_config('purple_color', foreground='#800080')  # Brown color code
        self.InputText.tag_config('blue_color', foreground='blue')  # Brown color code
        self.InputText.tag_config('BW_color', foreground=mainGUIFrame.rgb_to_hex(241,120,80))#call rgb_to_hex by class spacename   

###############################################
# UART Tx/Rx demo
# A simple Information Window
class InformWindow:
    def __init__(self,informStr):
        self.window = Tk()
        self.window.title("Information")
        self.window.geometry("220x60")
        label = Label(self.window, text=informStr)
        buttonOK = Button(self.window,text="OK",command=self.processButtonOK)
        label.pack(side = TOP)
        buttonOK.pack(side = BOTTOM)
        self.window.mainloop()

    def processButtonOK(self):
        self.window.destroy()


        
def create_win_row(switch_dict, ser_frame_instance, parent, item_info, default_color, row_index ):
    global label_2nd_var 
    global ENG_MODE_VAR
    global ENG_MODE
    global focus_out_flag
    global DIM_ID
    global READ_NUM
    is_updating = False
    label_trim_value = None
    label_iteration_num = []
    radio_button = None
    radio_button = []
    check_button= []
    label_2nd = []
    label_2nd_widget = []
    frame_radio = None
    saved_value = "" 
    focus_out_flag = False
    frame_row = Frame(parent)
    write_button = None 
    load_dim_id = DIM_ID
    #debug_print("load_dim_id is {}, DIM_ID is {}", load_dim_id, DIM_ID) 
    #frame_row.grid(row=row_index, column=0, sticky=W+E)
 # Add a button to the frame_row using grid
    button = Button(frame_row, text="Click me", command=lambda: print("Button clicked"))
    button.grid(row=0, column=0, sticky="w")  # Adjust row, column, and sticky options as needed

    # Configure row and column weights to make the grid expand
    frame_row.grid_rowconfigure(0, weight=1)
    frame_row.grid_columnconfigure(0, weight=1)

    #frame = tk.Frame(parent)
    item_text, default_value , write_permission, options_item = item_info
    check_var = IntVar()
    window_title = parent.title()


    if item_text[0:2] == "SW":                             
        check_button = Checkbutton(parent, text=item_text, variable=check_var,  
                                  command=lambda: update_dc_write_color(label_2nd, check_var), anchor='w', bg=default_color)
        # check_button = Checkbutton(parent, text=item_text, variable=check_var,  
    elif not item_text[0:2] == "SW":                             
        check_button = Checkbutton(parent, width=20,text=item_text, variable=check_var,  
                                  command=lambda: update_dc_write_color(label_2nd, check_var), anchor='w', bg=default_color)
     # Use sticky to ensure left alignment of check buttons
    check_button.grid(row=row_index, column=0, columnspan=3, padx=(0, 5), sticky='w')
    check_button.bind("<Enter>", lambda event, button=check_button: button_hover(event, button))
    check_button.bind("<Leave>", lambda event, button=check_button: button_leave(event, button))

    # Initially, all labels are created as disabled buttons
    label_2nd_var = StringVar()
    
    # Check if item_text is present in file_content
    if item_text[0:4] in file_content:
    # Extract the value from file_content
        debug_print("item_text[0:4] is {}", item_text[0:4])
        default_value = extract_value_from_file_content(item_text, file_content)
        default_value_int  = int(default_value, 16)
        if item_text[2:4] == '01':
            load_dim_id = default_value_int
            DIM_ID = default_value_int
            debug_print("DIM_ID is load_default_value_int is {}", DIM_ID)
        if item_text[2:4] == '00':
            READ_NUM = default_value
        label_2nd_var.set(default_value)
        label_2nd = Label(parent, textvariable=label_2nd_var, font=("Helvetica", 12), padx=2, pady=0, anchor='e', fg='RED', bg=default_color)
    else:
    # Use the default value if item_text is not in file_content
        label_2nd_var.set(default_value)
        label_2nd = Label(parent, textvariable=label_2nd_var, font=("Helvetica", 10), padx=2, pady=0, anchor='e', bg=default_color)

    label_2nd.grid(row=row_index, column=3, sticky='e')

    entry_var = StringVar(value=default_value)                                  
    entry = None # Placeholder for the Entry widget

    if write_permission == "WR" and not item_text[0:2] =="SW" and not item_text[0:2] == "OP":
        write_button = Button(parent, text="WRITE", state=NORMAL, command=lambda it=item_text, ev=entry_var, bg="#eeeeee",is_write=True:
            button_pressed(it, ev, write_button, is_write), bd=2, relief=GROOVE, activebackground=NAVY_BLUE, activeforeground="#ff0000",bg="#eeeeee")
        """
        ##################################################################################
        ###    Test ENG_MODE_VAR from cmd to enable write_button as NORMAL and DISABLED
        if window_title == 'win_otp' and not ENG_MODE_VAR.get() == 'write':
            write_button = Button(parent, text="WRITE", state=DISABLED, command=lambda it=item_text, ev=entry_var, bg=ORANGE,is_write=True:
                             button_pressed(it, ev, write_button, is_write), bd=5, relief=GROOVE, activebackground=BROWN, bg=ORANGE)
        else:                      
            write_button = Button(parent, text="WRITE", state=NORMAL, command=lambda it=item_text, ev=entry_var, bg=ORANGE,is_write=True:
                             button_pressed(it, ev, write_button, is_write), bd=5, relief=GROOVE, activebackground=BROWN, bg=ORANGE)
        ##################################################################################
        """

        # Bind the <Button-1> event to the button_pressed function
        write_button.bind("<Button-1>", lambda event, it=item_text, ev=entry_var, is_write=True:
                      button_pressed(it, ev, write_button, is_write))
        # Bind the <ButtonRelease-1> event to the button_released function
        write_button.bind("<ButtonRelease-1>", lambda event=None, it=item_text, ev=entry_var, is_write=True:
        #write_button.bind("<ButtonRelease-1>", lambda event, it=item_text, ev=label_2nd_var, is_write=True:
                      button_released(ser_frame_instance, row_index, it, label_2nd, entry, ev, write_button, is_write, parent, load_dim_id))
        write_button.grid(row=row_index, column=4, padx=(5,0), pady=(0,0)) 
        # Bind the hover and leave events
        write_button.bind("<Enter>", lambda event, button=write_button: button_hover(event, button))
        write_button.bind("<Leave>", lambda event, button=write_button: button_leave(event, button))
        # Update label_widgets_dict
        label_widgets_dict[row_index] = {"button_widget": write_button}
    elif not item_text[0:2] =="SW" and write_permission[0:1] == "W": # checkbutton also for selector OP                       
        write_button = Button(parent, text="WRITE", state=DISABLED, command=lambda it=item_text, ev=entry_var, bg="#eeeeee",is_write=True:
                             button_pressed(it, ev, write_button, is_write), bd=2, relief=GROOVE, activebackground=NAVY_BLUE,activeforeground=BROWN, bg="#eeeeee")
                             #button_pressed(it, ev, write_button, is_write), bd=5, relief=GROOVE, activebackground='grey', bg='light grey')
        write_button.grid(row=row_index, column=4, padx=(5,0), pady =(0,0))
        # Bind the hover and leave events
        write_button.bind("<Enter>", lambda event, button=write_button: button_hover(event, button))
        write_button.bind("<Leave>", lambda event, button=write_button: button_leave(event, button))
        label_widgets_dict[row_index] = {"button_widget": write_button}
    elif item_text[0:3] =="0xc": # create label width for current or voltage and iterate num                       
        #debug_print("it create trim_value label")
        label_trim_value = Label(parent, text=default_value, font=("Helvetica", 10), bg=default_color, padx=2, pady=0, anchor='e')
        label_trim_value.grid(row=row_index, column=4, sticky='e')
        label_iteration_num = Label(parent, text='00', font=("Helvetica", 10), bg=default_color, padx=0, pady=0, anchor='e', width = 3)
        label_iteration_num.grid(row=row_index, column=5, sticky='w')
        #debug_print("label_trim_value is {} iteration_num is {}",label_trim_value,label_iteration_num)
    ##########################################################################
    # Create switch button 
    ##########################################################################
    if item_text[0:2] =="SW":
        #debug_print("item_text: {} ",item_text)
        sw_nam = item_text[5:]
        app_switch = SwitchApp(parent, switch_name_attribute=sw_nam)
        # Use the dynamically selected color or default to None
        #switch_bg_color = selected_switch_color

        if sw_nam:
            app_switch.add_switch(row_index, 3, sw_nam, ser_frame_instance, 
            app_switch.img_eng_mode_on, app_switch.img_eng_mode_off, bg_color=default_color)
            #app_switch.add_switch(row_index, 3, sw_nam, ser_frame_instance, app_switch.img_eng_mode_on, app_switch.img_eng_mode_off)
            if window_title not in switch_dict:
                switch_dict[window_title] = {}
        # Store the app_switch instance in a dictionary
        #switch_dict[window_title][row_index] = app_switch # same as below
            #switch_dict[window_title][row_index] = app_switch # same as below
            switch_dict.setdefault(window_title, {})[row_index] = {
                "row": row_index,
                "app_switch_instance": app_switch,
            }
        else:
            debug_print("Invalid item_text: {} ", item_text)      

    ##########################################################################
    # Create radio button 
    ##########################################################################
    selected_option_var = StringVar()
    # print(f"item_text[0:4] is {item_text[0:4]}")
    selected_option_var.set(None)

    ########################################################### #
    def update_label(*args):
        global new_line
        try:
            selected_option = selected_option_var.get()
            #debug_print("selected_option is {}",selected_option)
            #label_2nd.config(text=selected_option)
            update_label_2nd_var(window_title, row_index, item_text, selected_option, "dark red")
            match item_text[0:4]:
                case 'OP01':
                    cmd_to_ser(new_line,ser_frame_instance, f"SET DC_SPEED {selected_option}")
                case 'OP02':
                    cmd_to_ser(new_line,ser_frame_instance, f"SET DC_low_count {selected_option}")
                case 'OP03':
                    cmd_to_ser(new_line,ser_frame_instance, f"SET DB_SPEED {selected_option}")
                case 'OP04':
                    cmd_to_ser(new_line,ser_frame_instance, f"SET DB_low_count {selected_option}")
                case 'OP05':
                    cmd_to_ser(new_line,ser_frame_instance, f"SET frame_rate {selected_option}")
                case 'OP06':
                    cmd_to_ser(new_line,ser_frame_instance, f"SET DC_reset_bit {selected_option}")
                case 'OP07':
                    cmd_to_ser(new_line,ser_frame_instance, f"SET DB_reset_bit {selected_option}")
                case 'OP08':
                    cmd_to_ser(new_line,ser_frame_instance, f"SET IDACG {selected_option}")
                case 'OP09':
                    cmd_to_ser(new_line,ser_frame_instance, f"SET test_osc_source {selected_option}")
                case 'OP10':
                    cmd_to_ser(new_line,ser_frame_instance, f"SET test_led_channel {selected_option}")
                case 'OP11':
                    cmd_to_ser(new_line,ser_frame_instance, f"SET pam_or_pwm_measure {selected_option}")
                case 'OP12':
                    cmd_to_ser(new_line,ser_frame_instance, f"SET ina228_exist {selected_option}")
        except Exception as e:
            debug_print("Error updating label: {}",e)

    # Attach the trace to the StringVar
    #selected_option_var.trace_add("write", update_label)

    if item_text[0:2] =="OP": # this is selector by radio button , created and occupy next row
        window_title = parent.title()

        # Check if there are old Radiobuttons and delete them
        if "radio_button_frame" in label_widgets_dict.get(window_title, {}).get(row_index, {}):
            old_frame = label_widgets_dict[window_title][row_index]["radio_button_frame"]
            old_frame.destroy()

        # Create a new frame for the Radiobuttons
        frame_radio = Frame(parent, bg=default_color)
        frame_radio.grid(row=row_index +1 , column=0, columnspan=len(options_item))

        # Ensure the key exists in the dictionary
        label_widgets_dict.setdefault(window_title, {}).setdefault(row_index, {})
        # Initialize an empty dictionary to store radio buttons
        label_widgets_dict[window_title][row_index]["radio_button"] = []
        #label_widgets_dict[window_title][row_index].setdefault("radio_button", [])

        for idx, option in enumerate(options_item):
            # debug_print("here 2")
            radio_button = Radiobutton(frame_radio, text=option, bg=default_color, variable=selected_option_var, value=option, command=update_label)
            radio_button.grid(row=0, column= idx, sticky=W)
            label_widgets_dict[window_title][row_index]["radio_button"].append(radio_button) 
            # Ensure the key exists in the dictionary

        #label_widgets_dict.setdefault(window_title, {})
        #label_widgets_dict[window_title].setdefault(row_index, {})

        label_widgets_dict[window_title][row_index]["radio_button_frame"] = frame_radio 

        row_index = row_index + 1            
        ##############################################################

    # Create a "READ" button for each row
    elif not item_text[0] == "S" and ( write_permission[0:1] == "R" or write_permission[1:2] == "R" ):                            
        read_button = Button(parent, text="READ", command=lambda it=item_text, ev=entry_var, is_write=False:
                            button_pressed(it, ev, read_button, is_write), bd=2, relief=GROOVE, activebackground='#eeeeee', bg='#eeeeee')
        # Bind the <Button-1> event to the button_pressed function
        read_button.bind("<Button-1>", lambda event, it=item_text, ev=entry_var,is_write=False:
                     button_pressed(it, ev, read_button, is_write))
        # Bind the <ButtonRelease-1> event to the button_released function
        read_button.bind("<ButtonRelease-1>", lambda event, it=item_text, ev=entry_var, is_write=False:
                     button_released(ser_frame_instance, row_index,it, label_2nd, entry, ev, read_button, is_write, parent, load_dim_id))
        read_button.grid(row=row_index, column=5)

        # Bind the hover and leave events
        read_button.bind("<Enter>", lambda event, button=read_button: button_hover(event, button))
        read_button.bind("<Leave>", lambda event, button=read_button: button_leave(event, button))

    ########################################################### #
    def on_entry_changed(event):
        global focus_out_flag
        focus_out_flag = True
        debug_print("on_entry_changed")
    ########################################################### #
    def label_2nd_clicked():
        nonlocal entry, label_2nd, saved_value # so special word : nonlocal frank NOTE

        # Save the current label value
        saved_value = label_2nd.cget("text")

        # Switch to Entry when label_2nd is clicked
        #entry_var = tk.StringVar(value=default_value)
        if item_text[0:3] =="0xc":
            entry_width = 7
        else:            
            entry_width = 2
        # print(f"entry_width is {entry_width}, item_text[0:4] is {item_text[0:4]}")
        entry = Entry(parent, 
                      width=entry_width, 
                      justify='right', 
                      textvariable=entry_var, font=("Helvetica", 10)
                      )
        entry.icursor(END) # important , let cursor sit on the right of entry 
        entry.grid(row=row_index, column=3, padx=(0, 0), pady=(0,0), sticky='e')
        ########################################################### #
        def entry_focus_out(event):
            nonlocal is_updating, entry, saved_value, label_2nd, parent, row_index, frame_radio, item_text
            global label_2nd_var
            global focus_out_flag
            global file_content
            if focus_out_flag:
                debug_print("Entry focus out")
                focus_out_flag = False

            if not is_updating:
                is_updating = True
                # Switch back to Label_2nd when finishing editing
                value = entry_var.get()
                # print(f" 320 entry_focus_out ")
                if not value.strip(): # check it empty
                    # print(f" 322 entry_focus_out ")
                    value = saved_value
                if not is_valid_hex(value): 
                    # print(f"invalid value {value}")
                    show_warning_message("Invalid Value", "Please enter a valid hex value between '00' to 'ff'")
                    entry.grid_remove()
                    entry_var.set("")
                else:                    
                    entry.grid_remove()
                    # Append new lines
                    if not value == saved_value:
                        file_content += f"{item_text[0:4]} {value}\n"
                    label_2nd.config(text=value)
                    label_2nd_var.set(value)
                    window_title = parent.title()
                    text_color = "blue"
                    update_label_2nd_var(window_title, row_index, item_text, value, text_color)
                    # print(f" 323 entry_focus_out, value is {value} ")
                    label_2nd.grid()
                    print(f"file_content is {file_content}\n")

                is_updating = False

        # Bind focus out event to switch back to Label_2nd
        entry.bind("<FocusOut>", entry_focus_out)
        entry.bind("<KeyRelease>", on_entry_changed)
        entry.focus_set()  # Set focus to the entry widget

    # Bind label_2nd click event to switch to Entry
    if not item_text[0:2] =="OP": # only not OP , not SW, the label_2nd_clicked will take effect 
        label_2nd.bind("<Button-1>", lambda event: label_2nd_clicked())

    # Function to update the text of a specific Entry widget in a row for a specific window
    #################################################################################
    # Store the entry_var and Entry widget in the dictionary
    #if "win_db_wr" not in label_widgets_dict:
    #    label_widgets_dict[parent.title()] = {}
    #label_widgets_dict[parent.title()][row_index] = {
    window_title = parent.title()
    if window_title not in label_widgets_dict:
        label_widgets_dict[window_title] = {}


    label_widgets_dict[window_title][row_index] = {
        "check_button": check_button,
        "label_2nd_var": label_2nd_var,
        "label_2nd_widget": label_2nd,
        "radio_button": [],
        #"radio_button": [radio_button_0, radio_button_1, radio_button_2, radio_button_3, radio_button_4],
        "radio_button_frame": frame_radio,
        "label_trim_value": label_trim_value,
        "label_iteration_num": label_iteration_num,
        "button_widget": write_button
    }            

    # print(f"282 keys in label_widgets_dict: {label_widgets_dict.keys()}, row is {row_index}") 
    return switch_dict ,  check_button, label_2nd, check_var, entry, row_index, frame_row

##################### end of gui class ################

# Usage example
#window_title = 'example_window'
#label_2nd_var_string = get_label_2nd_var_values(window_title)
#print(label_2nd_var_string)


def button_hover(event, button):
    if button["state"] == NORMAL:  # Check if the button is enabled
        button.config(fg="cyan")

def button_leave(event, button):
    if button["state"] == NORMAL:
        button.config(fg="black")

def button_pressed(item_text, entry_var, button, is_write):
    # Use get() to retrieve the actual string value from the StringVar
    # print(f"{button['text']} button pressed for {item_text}: {entry_var.get()}")
    # Change background color to grey
    button.config(bg='light grey')

def button_released(ser_frame_instance,row_index,item_text,label_2nd_widget, entry, entry_var, button, is_write, win_name, dim_id):
    global DIM_ID
    global READ_NUM # this is the read dc byte count number - 1 
    global READ_NUM_INT
    global START_REG_ADDR
    global READ_ROW_NUM
    global HT2_NUM
    action = {}
    win_name_text = win_name.title()
    result = win_name_text[5:6]
    ##########################################
    # when release WRITE button 
    ##########################################
    if is_write == True and not result == "r" and not result == "e": # r is trim window 
        if result == "c" or result== "t" : # this is dc or otp window
            action = "CW"
            if item_text[2:4] == "01":
                dim_id_input = dim_id
                dim_id = 0x00 
                first_byte = 0x00
            elif not item_text[2:4] == "00": 
                first_byte = (DIM_ID << 1) | 0x00 
        elif result == "b": # this is db window            
            action = "BW"
            if item_text[2:4] == "01":
                dim_id_input = dim_id
                dim_id = 0x00 
                first_byte = 0x00
                debug_print("here 1")
            else : 
                first_byte = (DIM_ID << 1) | 0x00 
        else :            
            first_byte = (DIM_ID << 1) | 0x00 
        length = 0x00
        value = entry_var.get()
        value_int  = int(value, 16)
        if not is_valid_hex(value): 
            show_warning_message("Invalid Value", "Please enter a valid hex value between '00' to 'ff'")
        elif  item_text[0:4] == "0x00":   
            value_int  = int(value, 16) - 1
            print(f" READ_NUM : {READ_NUM}")
            READ_NUM_INT = value_int
            print(f" READ_NUM_INT : {READ_NUM_INT}")
            READ_NUM = hex(value_int)[2:].zfill(len(value))
            print(f" READ_NUM : {READ_NUM}")
        elif ( action == "CW" or action == "BW")  and item_text[2:4] == "01" and not value == "00":
            string = '{:02X} {} {:02} {:02}'.format(first_byte, item_text[2:4], length, '00') # this is to change dimm_id cmd.
            crc = crc8(f"{string}") 
            thread = threading.Thread(target=cmd_to_ser(new_line,ser_frame_instance, f"{action} {string} {crc}"))
            thread.start()
            dim_id = value
            DIM_ID = value_int
            debug_print("DIM_ID is {}",DIM_ID)
            dis_dbo_dim_id = value_int | 0x80
            value = format(dis_dbo_dim_id, '02X')
            string = '{:02X} {} {:02} {:02}'.format(first_byte, item_text[2:4], length, value) # this is to change dimm_id cmd.
            crc = crc8(f"{string}") 
            thread.join()
            cmd_to_ser_continous(ser_frame_instance, f"{action} {string} {crc}")
            #cmd_to_ser(new_line,ser_frame_instance, f"{action} {string} {crc}")
        else:      
            string = '{:02X} {} {:02} {:02}'.format(first_byte, item_text[2:4], length, value) # this is to change dimm_id cmd.
            if item_text[2:4] == "aa": # read all of data from reg 10 to reg 0x27, startting_row_index = 3
                label_2nd_var_string = get_label_2nd_var_values('win_db_wr') 
                #debug_print("label_2nd_var_string is {}",label_2nd_var_string)
                string = '{:02X} {} {:02} {:02}'.format(first_byte, 10, 17, label_2nd_var_string) # this is to change dimm_id cmd.
            crc = crc8(f"{string}") 
            cmd_to_ser(new_line,ser_frame_instance, f"{action} {string} {crc}" )
        # Trigger focus out behavior for the entry
        if entry:
            #entry.focus_force()
            label_2nd_widget.focus_set()
            debug_print("focus_force Entry is", entry)
        else:
            debug_print("Entry is None")

    ##########################################
    # when release READ button and is_write is false 
    ##########################################
    elif not is_write and not result == "e":         
        if result == "c":  # this is dc or otp window
            action = "CRC" # read register cmd is always 5 byte to send, add the row_index to be 6 byte
        elif result == "b":  # this is dc or otp window
            action = "CRB" # read register cmd is always 5 byte to send, add the row_index to be 6 byte
        elif result == "t":  # this is dc or otp window
            action = "CRT" # read register cmd is always 5 byte to send, add the row_index to be 6 byte
            
        first_byte = (DIM_ID << 1) | 0x01 
        print(f"READ_NUM: {READ_NUM}")
        START_REG_ADDR = item_text[2:4] 
        READ_ROW_NUM = row_index 
        print(f" START_REG_ADDR: {item_text[2:4]}, READ_ROW_NUM: {READ_ROW_NUM}")
        print(f" START_REG_ADDR: {START_REG_ADDR}")
        string = '{:02X} {} {:02}'.format(first_byte, item_text[2:4], READ_NUM) # this is to change dimm_id cmd.
        #string = '{:02X} {} {:02}'.format(first_byte, item_text[2:4], length) # this is to change dimm_id cmd.
        if item_text[2:4] == "aa": # read all of data from reg 10 to reg 0x27, startting_row_index = 3
            string = '{:02X} {} {:02}'.format(first_byte, 10, 17) # hex 17 is decimal 24 -1 = 23.
            row_index = 3
        crc = crc8(f"{string}") 
        cmd_to_ser(new_line,ser_frame_instance, f"{action} {string} {crc} {row_index}")
        # should wait to receive from ser, then trun to light grey
        #wait_ser()
    elif is_write == True and result == "r": # win_trim and write  
        value = entry_var.get()
        if item_text[2:4] == 'v2':
            action = "SET"
            string = 'v2 '+ value
        cmd_to_ser(new_line,ser_frame_instance, f"{action} {string}")
            #update_label_trim_var("win_trim", 7, "011110", 6, 'blue') # correct method to call this function, 7 is row 7 from 1
    elif is_write == True and result == "e": # win_trim and write  
        value = entry_var.get()
        action = "SET"
        if item_text[2:4] == 't1':
            string = 'TI_ADC_CONFIG_H '+ value
        elif item_text[2:4] == 't2':
            string = 'TI_ADC_CONFIG_L '+ value
        elif item_text[2:4] == 't3':
            string = 'HT2_NUM '+ value
            HT2_NUM = value
        cmd_to_ser(new_line,ser_frame_instance, f"{action} {string}")
    elif not is_write and result == "e": # win_trim and write  
        action = "GET"
        if item_text[2:4] == 't1':
            string = 'TI_ADC_CONFIG_H'
        elif item_text[2:4] == 't2':
            string = 'TI_ADC_CONFIG_L'
        elif item_text[2:4] == 't3':
            string = 'HT2_NUM'
        cmd_to_ser(new_line,ser_frame_instance, f"{action} {string}")
    #elif not is_write and result == "e":         
    # Reset background color to USER SETTING

    button.config(bg=MALACHITE_GREEN)  # Change this to the desired default color

import sys
#sys.path.append('/Users/at891/source/prgtst')
sys.path.append('at891/source/prgtst/SerialMonitor_main')

#from . import frank  
#from SerialMonitor_main import frank  
def on_menu_item_hover(event):
    # Change the foreground color only if the menu item is not disabled
    if event.widget["state"] != DISABLED:
        event.widget.config(fg="blue")

def on_menu_item_leave(event):
    # Reset the foreground color when leaving the menu item
    event.widget.config(fg="black")

def on_menu_enter(event):
    #if not event.widget.instate(['disabled']):
    event.widget.configure(background="#FFBFFF")
        #event.widget.configure(background="#00BFFF")
    #menu_bar.entryconfig(background="deep sky blue")

def on_menu_leave(event):
    #if not event.widget.instate(['disabled']):
    event.widget.configure(background="#FFBFFF")
    #menu_bar.entryconfig(event.widget.name, background="")

# Define our switch function
class SwitchApp:
    def __init__(self, parent, switch_name_attribute):
        self.parent = parent
        #self.parent.title("Switch Example")
        self.switch_name_attribute = switch_name_attribute
        self.buttons = {}  # Dictionary to store buttons and their corresponding images

        # Define image pairs for different modes
        self.img_eng_mode_on = PhotoImage(file="images/img_eng_mode_on.png")
        self.img_eng_mode_off = PhotoImage(file="images/img_eng_mode_off.png")
        self.img_vpp_on = PhotoImage(file="images/img_eng_mode_on.png")
        self.img_vpp_off = PhotoImage(file="images/img_eng_mode_off.png")

    def add_switch(self, row, column, button_name, class_instance, on_image, off_image, bg_color):
        #debug_print("special button_name is {}",button_name)
        button = Button(
            self.parent, 
            command=lambda: self.toggle_mode(button_name, class_instance), 
            image=off_image,
            borderwidth = 0,
            bg = bg_color,
            highlightthickness=0,
        )
        button.grid( row=row, column=column)
        self.buttons[button_name] = {"button": button, "on_image": on_image, "off_image": off_image, "state": False, "row": row}

    def toggle_mode(self, sw_nam, class_instance):
        #debug_print("class_instance is {}", class_instance)
        global ENG_MODE, AUTO_SEND_CMD, VPP_TO_10V
        global VDD_TO_5V, CHOP_FREQ, ON_LINE_CALIBR, LOCK, IS_PAOWANG
        global VREF_TRIM, CURRENT_TRIM
        global FREQ_COARSE_TRIM,FREQ_FINETUNE_TRIM,PRINT_MESSAGE 
        global EXTCLK_ENABLE, DIV256_ENABLE, AUTO_SEARCH_FREQ 
        global DC_RESET, DB_REPEAT
        global new_line
        global GET_LED_CURRENT 
        #debug_print("DIM_ID is : {}", DIM_ID)
        sw_dat = self.buttons.get(sw_nam)
        if sw_dat:
            sw_dat["state"] = not sw_dat["state"]
            if sw_dat["state"]:
                sw_dat["button"].config(image=sw_dat["on_image"])
                match sw_nam:
                    case'ENGINEER MODE':
                        crc = crc8('01 ee 00')
                        cmd_to_ser(new_line,class_instance, f"CW 01 ee 00 {crc} # engineer mode enter")
                        ENG_MODE = True

                    case 'AUTO SEND CMD':
                        cmd_to_ser(new_line,class_instance, f"SW 02 1 #enable AUTO SEND CMD")
                        AUTO_SEND_CMD = True

                    case 'AUTO SEARCH FREQ':
                        cmd_to_ser(new_line,class_instance, f"SW AUTO_SEARCH_FREQ 1")
                        AUTO_SEARCH_FREQ = True

                    case 'VPP to 10V':
                        cmd_to_ser(new_line,class_instance, f"SW 03 1 # SET Vpp to 10V before Program otp")
                        VPP_TO_10V = True
                    case 'VDD to 5V':
                        cmd_to_ser(new_line,class_instance, f"SW 04 1 # SET VDD to 5V before Program otp")
                        VDD_TO_5V = True
                    case 'CHOP FREQ TEST': 
                        cmd_to_ser(new_line,class_instance, f"SW 06 1 # ENABLE CHOP FREQ TEST")
                        CHOP_FREQ = True
                    case 'on_line_Calibr': 
                        cmd_to_ser(new_line,class_instance, f"SW 07 1 # ENABLE starting on_line calibration")
                        ON_LINE_CALIBR = True
                    case 'LOCK': 
                        cmd_to_ser(new_line,class_instance, f"SW 08 1 # LOCK the register write")
                        LOCK = False
                    case 'IS_PAOWANG': 
                        cmd_to_ser(new_line,class_instance, f"SW 09 1 # Switch to PAOWANG design")
                        IS_PAOWANG = True
                    case 'VREF TRIM': 
                        #crc = crc8('01 ee 11')
                        #cmd_to_ser_continous(class_instance, f"CW 01 ee 11 {crc} # ENABLE Vref Trim process\n")
                        #cmd_to_ser_continous(class_instance, f"SW Vref_trim 1 # ENABLE Vref Trim process")
                        cmd_to_ser(new_line, class_instance, f"SW Vref_trim 1 # ENABLE Vref Trim process")
                        VREF_TRIM = True
                    case 'AVDDL from DBO': 
                        crc = crc8('01 ee 11')
                        cmd_to_ser(new_line, class_instance, f"CW 01 ee 11 {crc}")
                        self.parent.after(2000, lambda: self.switch_off(sw_nam))
                    case 'CURRENT TRIM': 
                        #debug_print("sw_nam is {} ",sw_nam)
                        cmd_to_ser(new_line, class_instance, f"SW CURRENT_TRIM 1")
                        CURRENT_TRIM = True
                    case 'FREQ COARSE TRIM': 
                        cmd_to_ser(new_line, class_instance, f"SW TRIM_COARSE_FREQ 1 # ENABLE COARSE FREQ Trim process")
                        FREQ_COARSE_TRIM = True
                    case 'DB Repeat': 
                        cmd_to_ser(new_line, class_instance, f"SW DB_REPEAT 1 # ENABLE repeat DB data frame")
                        DB_REPEAT = True
                    case 'DC RESET': 
                        cmd_to_ser(new_line, class_instance, f"RESET_DC")
                        self.parent.after(1000, lambda: self.switch_off(sw_nam))
                        DC_RESET = True
                    case 'DB RESET': 
                        cmd_to_ser(new_line, class_instance, f"RESET_DB")
                        self.parent.after(1000, lambda: self.switch_off(sw_nam))
                        DC_RESET = True
                    case 'FREQ FINETUNE TRIM': 
                        cmd_to_ser(new_line, class_instance, f"SW TRIM_FINE_FREQ 1 # ENABLE FINETUNE FREQ Trim process")
                        FREQ_FINETUNE_TRIM = True
                    case 'PRINT_MESSAGE': 
                        cmd_to_ser(new_line, class_instance, f"SW 15 1 # Switch to ENABLE PRINT MESSAGE")
                        PRINT_MESSAGE = True
                    case 'GET_LED_CURRENT': 
                        cmd_to_ser(new_line, class_instance, f"SW GET_LED_CURRENT 1")
                        self.parent.after(1000, lambda: self.switch_off(sw_nam))
                        GET_LED_CURRENT = True
                    case 'NEW LINE ON CMD': 
                        new_line = True
                        cmd_to_ser(new_line, class_instance, f"# new_line is True")
                    case 'ExtClk Enable': 
                        cmd_to_ser(new_line, class_instance, f"SW extclk 1")
                        EXTCLK_ENABLE = True
                    case 'OSC div256 Enable': 
                        DIV256_ENABLE = True         
                        string = '01 ee 29'
                        crc = crc8(f"{string}") 
                        cmd_to_ser(new_line, class_instance, f"CW {string} {crc}  #OSC div256 ENABLE")
            else:
                sw_dat["button"].config(image=sw_dat["off_image"])
                match sw_nam:
                    case'ENGINEER MODE':
                        crc = crc8('01 ef 00')
                        cmd_to_ser(new_line, class_instance, f"CW 01 ef 00 {crc} # engineer mode leave")
                        ENG_MODE = False
                    case 'AUTO SEND CMD':
                        cmd_to_ser(new_line, class_instance, f"SW 02 0 # DISABLE AUTO SEND CMD")
                        AUTO_SEND_CMD = False

                    case 'AUTO SEARCH FREQ':
                        cmd_to_ser(new_line,class_instance, f"SW AUTO_SEARCH_FREQ 0")
                        AUTO_SEARCH_FREQ = False

                    case 'VPP to 10V':
                        cmd_to_ser(new_line, class_instance, f"SW 03 0 # SET Vpp to 5V after program otp done")
                        VPP_TO_10V = False
                    case 'VDD to 5V':
                        cmd_to_ser(new_line, class_instance, f"SW 04 0 # SET VDD to 3.3V after program otp done")
                        VDD_TO_5V = False
                    case 'CHOP_FREQ_TEST': 
                        cmd_to_ser(new_line, class_instance, f"SW 06 0 # DISABLE CHOP FREQ TEST")
                        CHOP_FREQ = False
                    case 'on_line_Calibr': 
                        cmd_to_ser(new_line, class_instance, f"SW 07 0 # DISABLE starting on_line calibration")
                        ON_LINE_CALIBR = False
                    case 'LOCK': 
                        cmd_to_ser(new_line, class_instance, f"SW 08 0 # UNLOCK the register write")
                        LOCK = False
                    case 'NEW LINE ON CMD': 
                        new_line = False
                        cmd_to_ser(new_line, class_instance, f"# new_line is False")
                    case 'IS_PAOWANG': 
                        cmd_to_ser(new_line, class_instance, f"SW 09 0")
                        IS_PAOWANG = False
                    case 'VREF TRIM': 
                        cmd_to_ser(new_line, class_instance, f"SW Vref_trim 0")
                        VREF_TRIM = False
                    case 'CURRENT TRIM': 
                        cmd_to_ser(new_line, class_instance, f"SW CURRENT_TRIM 0")
                        CURRENT_TRIM = False
                    case 'FREQ COARSE TRIM': 
                        cmd_to_ser(new_line, class_instance, f"SW TRIM_COARSE_FREQ 0")
                        FREQ_COARSE_TRIM = False
                    case 'FREQ FINETUNE TRIM': 
                        cmd_to_ser(new_line, class_instance, f"SW TRIM_FINE_FREQ 0")
                        FREQ_FINETUNE_TRIM = False
                    case 'DB Repeat': 
                        cmd_to_ser(new_line, class_instance, f"SW DB_REPEAT 0")
                        DB_REPEAT = False
                    case 'GET_LED_CURRENT': 
                        cmd_to_ser(new_line, class_instance, f"SW GET_LED_CURRENT 0")
                        GET_LED_CURRENT = False
                    case 'PRINT_MESSAGE': 
                        cmd_to_ser(new_line, class_instance, f"SW 15 0")
                        PRINT_MESSAGE = False
                    case 'ExtClk Enable': 
                        cmd_to_ser(new_line, class_instance, f"SW extclk 0")
                        EXTCLK_ENABLE = False
                    case 'OSC div256 Enable': 
                        cmd_to_ser(new_line, class_instance, f"ENG osc/256 0")
                        DIV256_ENABLE = False

    def switch_off(self, sw_nam):
        sw_dat = self.buttons.get(sw_nam)
        if sw_dat and sw_dat["state"]:
            sw_dat["state"] = False
            sw_dat["button"].config(image=sw_dat["off_image"])


####################  end of snippet for create switch button ########################
def button_pressed(item_text, entry_var, button, is_write):
    # Use get() to retrieve the actual string value from the StringVar
    # print(f"{button['text']} button pressed for {item_text}: {entry_var.get()}")
    # Change background color to grey
    button.config(bg='light grey')

####################  end of snippet for create switch button ########################
def button_save():
    from ..frank import btn_1_1, root 
    btn_1_1.config(text="Save it!")
    root.update_idletasks()  # Update the display
    root.after(1000, lambda: btn_1_1.config(text="Save"))

def write_button_clicked(item_text, entry_var):
    # Replace this with your logic to handle the write button click
    temp_var = entry_var.get()
    debug_print("Write button clicked for {}: {}",item_text, temp_var)

def read_button_clicked(item_text, entry_var):
    # Replace this with your logic to handle the read button click
    temp_var = entry_var.get()
    debug_print("Read button clicked for {}: {}",item_text, temp_var)

def button_hover(event, button):
    if button["state"] == NORMAL:  # Check if the button is enabled
        button.config(fg="cyan")

def button_leave(event, button):
    if button["state"] == NORMAL:
        button.config(fg="black")
